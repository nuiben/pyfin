import openpyxl
import os

os.chdir('C:\\Users\\Ben\\nuiben\\pyfin\\Excel\\purchases')

master_data = openpyxl.load_workbook("master_sheet.xlsx")
daily_data = openpyxl.load_workbook("daily_sheet.xlsx")

master_sheet = master_data['master']
daily_sheet = daily_data['Sheet1']

# get row count for daily sheet
is_data = True
daily_row_count = 1

while is_data:
    daily_row_count += 1
    data = daily_sheet.cell(row=daily_row_count, column=1).value
    # print(data)
    if data == None:
        daily_row_count -= 2
        is_data = False
# print(daily_row_count)

# get row count for master sheet
# we should definitely make this a function instead of repeating work - bp
